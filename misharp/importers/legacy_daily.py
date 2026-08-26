from __future__ import annotations

import re
from datetime import date, datetime
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from ..db import session_scope
from ..models import DailyCondition
from ..repositories import finish_sync_run, start_sync_run

SHEET_RE = re.compile(r"(?P<year>20\d{2})년\s*(?P<month>\d{1,2})월")

ALIASES = {
    "paid_amount": ["실결제"],
    "ad_cost": ["일별광고비", "광고비"],
    "purchase_count": ["구매건수"],
    "avg_order_value": ["객단가"],
    "conversion_rate": ["전환율"],
    "visitors": ["전체방문"],
    "pageviews": ["페이지뷰"],
    "search_visits": ["검색방문"],
    "ad_visits": ["광고유입"],
    "bookmark_visits": ["웹북마크"],
    "app_installs": ["앱설치수", "앱 설치수"],
    "app_unique_visits": ["앱순방문", "앱 순방문"],
    "shipping_count": ["택배수량", "택배 수량"],
    "member_signups": ["회원가입", "회원 가입"],
}


def norm(value) -> str:
    return re.sub(r"\s+", "", str(value or "")).strip()


def _as_float(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = re.sub(r"[^0-9.\-]", "", str(value))
    try:
        return float(text) if text else None
    except ValueError:
        return None


def _as_int(value):
    n = _as_float(value)
    return int(round(n)) if n is not None else None


def _normalize_conversion(value):
    n = _as_float(value)
    if n is None:
        return None
    return n * 100 if 0 < abs(n) <= 1 else n


def _parse_day(value, year: int, month: int, epoch) -> date | None:
    if value in (None, ""):
        return None
    d = None
    if isinstance(value, datetime):
        d = value.date()
    elif isinstance(value, date):
        d = value
    elif isinstance(value, (int, float)):
        try:
            d = from_excel(value, epoch=epoch).date()
        except Exception:
            return None
    else:
        text = re.sub(r"\([^)]*\)", "", str(value)).strip()
        for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d"):
            try:
                d = datetime.strptime(text, fmt).date()
                break
            except ValueError:
                pass
        if d is None and text.isdigit():
            try:
                d = date(year, month, int(text))
            except ValueError:
                return None
    return d if d and d.year == year and d.month == month else None


def _parse_workbook(file_bytes: bytes) -> tuple[list[dict], list[str]]:
    wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    records: list[dict] = []
    notes: list[str] = []

    for ws in wb.worksheets:
        m = SHEET_RE.search(ws.title)
        if not m:
            continue
        year, month = int(m.group("year")), int(m.group("month"))
        rows = list(ws.iter_rows(min_row=1, max_row=60, min_col=1, max_col=80, values_only=True))
        header_idx = None
        headers: list[str] = []
        for idx, row in enumerate(rows):
            vals = [norm(v) for v in row]
            if "월일" in vals and "실결제" in vals:
                header_idx, headers = idx, vals
                break
        if header_idx is None:
            notes.append(f"{ws.title}: 헤더 미탐지")
            continue

        date_col = headers.index("월일")
        colmap: dict[str, int] = {}
        for field, aliases in ALIASES.items():
            wanted = {norm(x) for x in aliases}
            for i, h in enumerate(headers):
                if h in wanted:
                    colmap[field] = i
                    break

        for row in rows[header_idx + 1:]:
            d = _parse_day(row[date_col] if date_col < len(row) else None, year, month, wb.epoch)
            if not d:
                continue
            values = {}
            for field, col in colmap.items():
                raw = row[col] if col < len(row) else None
                if field == "conversion_rate":
                    v = _normalize_conversion(raw)
                elif field in {"purchase_count", "visitors", "pageviews", "search_visits", "ad_visits", "bookmark_visits", "app_installs", "app_unique_visits", "shipping_count", "member_signups"}:
                    v = _as_int(raw)
                else:
                    v = _as_float(raw)
                if v is not None:
                    values[field] = v
            if values:
                records.append({"date": d, **values})
    wb.close()
    return records, notes


def preview_legacy_daily(file_bytes: bytes) -> tuple[pd.DataFrame, list[str]]:
    records, notes = _parse_workbook(file_bytes)
    if not records:
        raise RuntimeError("과거 일일보고에서 월별 일자 데이터를 찾지 못했습니다.")
    df = pd.DataFrame(records).sort_values("date")
    preview = pd.DataFrame([
        {
            "시작일": df["date"].min(),
            "종료일": df["date"].max(),
            "일자수": int(df["date"].nunique()),
            "실결제 포함일": int(df["paid_amount"].notna().sum()) if "paid_amount" in df else 0,
            "방문 포함일": int(df["visitors"].notna().sum()) if "visitors" in df else 0,
        }
    ])
    return preview, notes


def import_legacy_daily(file_bytes: bytes, file_name: str) -> tuple[int, int, list[str]]:
    """과거 일일보고를 DB에 채운다.

    기존 Cafe24/Google/iApps 등 최신 수집값은 절대 덮어쓰지 않고,
    DB에서 비어 있는 컬럼만 과거 엑셀 값으로 보충한다.
    """
    records, notes = _parse_workbook(file_bytes)
    if not records:
        raise RuntimeError("과거 일일보고에서 적재할 데이터를 찾지 못했습니다.")

    inserted_days = 0
    filled_fields = 0
    with session_scope() as db:
        run = start_sync_run(db, "legacy_excel")
        try:
            for rec in records:
                d = rec.pop("date")
                row = db.get(DailyCondition, d)
                if row is None:
                    row = DailyCondition(date=d)
                    db.add(row)
                    inserted_days += 1
                for field, value in rec.items():
                    if getattr(row, field, None) is None and value is not None:
                        setattr(row, field, value)
                        filled_fields += 1
                sources = dict(row.sources or {})
                sources["legacy"] = file_name
                row.sources = sources
            finish_sync_run(
                db,
                run,
                "success",
                rows_written=len(records),
                message=f"{file_name} · {filled_fields}개 빈 필드 보충",
            )
        except Exception as exc:
            finish_sync_run(db, run, "failed", 0, str(exc))
            raise
    return len(records), filled_fields, notes
