from __future__ import annotations

import argparse
import re
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from misharp.db import init_db, session_scope
from misharp.models import DailyCondition
from misharp.repositories import upsert_daily

SHEET_RE = re.compile(r"(?P<year>20\d{2})년\s*(?P<month>\d{1,2})월")

ALIASES = {
    "paid_amount": ["실결제"],
    "ad_cost": ["일별광고비", "광고비"],
    "purchase_count": ["구매건수"],
    "avg_order_value": ["객단가"],
    "conversion_rate": ["전환율"],
    "visitors": ["전체방문"],
    "pageviews": ["페이지뷰"],
    "search_visits": ["검색방문"],
    "ad_visits": ["광고유입"],
    "bookmark_visits": ["웹북마크"],
    "app_installs": ["앱설치수", "앱 설치수"],
    "app_unique_visits": ["앱순방문", "앱 순방문"],
    "shipping_count": ["택배수량", "택배 수량"],
    "member_signups": ["회원가입", "회원 가입"],
}


def norm(value) -> str:
    return re.sub(r"\s+", "", str(value or "")).strip()


def parse_day(value, year: int, month: int, epoch) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        d = value.date()
    elif isinstance(value, date):
        d = value
    elif isinstance(value, (int, float)):
        try:
            d = from_excel(value, epoch=epoch).date()
        except Exception:
            return None
    else:
        text = re.sub(r"\([^)]*\)", "", str(value)).strip()
        for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d"):
            try:
                d = datetime.strptime(text, fmt).date()
                break
            except ValueError:
                d = None
        if d is None and text.isdigit():
            try:
                d = date(year, month, int(text))
            except ValueError:
                return None
    return d if d is not None and d.year == year and d.month == month else None


def as_float(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = re.sub(r"[^0-9.\-]", "", str(value))
    try:
        return float(text) if text else None
    except ValueError:
        return None


def as_int(value):
    number = as_float(value)
    return int(round(number)) if number is not None else None


def normalize_conversion(value):
    n = as_float(value)
    if n is None:
        return None
    # 과거 시트는 0.0236, 신형 시트는 2.36 형식이 혼재한다.
    return n * 100 if 0 < abs(n) <= 1 else n


def find_header(rows):
    for idx, row in enumerate(rows):
        values = [norm(v) for v in row]
        if "월일" in values and "실결제" in values:
            return idx, values
    return None, []

def import_file(path: Path, dry_run: bool = False) -> tuple[int, list[str]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    total = 0
    notes: list[str] = []

    for ws in wb.worksheets:
        m = SHEET_RE.search(ws.title)
        if not m:
            continue
        year, month = int(m.group("year")), int(m.group("month"))
        rows = list(ws.iter_rows(min_row=1, max_row=60, min_col=1, max_col=80, values_only=True))
        header_idx, headers = find_header(rows)
        if header_idx is None:
            notes.append(f"{ws.title}: 헤더 미탐지")
            continue

        date_col = headers.index("월일")
        colmap: dict[str, int] = {}
        for field, aliases in ALIASES.items():
            normalized_aliases = {norm(x) for x in aliases}
            for i, h in enumerate(headers):
                if h in normalized_aliases:
                    colmap[field] = i
                    break

        for row in rows[header_idx + 1:]:
            d = parse_day(row[date_col] if date_col < len(row) else None, year, month, wb.epoch)
            if not d:
                continue
            values = {}
            for field, col in colmap.items():
                raw = row[col] if col < len(row) else None
                if field == "conversion_rate":
                    values[field] = normalize_conversion(raw)
                elif field in {"purchase_count", "visitors", "pageviews", "search_visits", "ad_visits", "bookmark_visits", "app_installs", "app_unique_visits", "shipping_count", "member_signups"}:
                    values[field] = as_int(raw)
                else:
                    values[field] = as_float(raw)
            values = {k: v for k, v in values.items() if v is not None}
            if not values:
                continue
            total += 1
            if dry_run:
                continue
            with session_scope() as db:
                existing = db.get(DailyCondition, d)
                sources = dict(existing.sources or {}) if existing else {}
                sources["legacy"] = path.name
                upsert_daily(db, d, **values, sources=sources)

    wb.close()
    return total, notes


def main() -> None:
    parser = argparse.ArgumentParser(description="미샵 과거 월별 일일보고 엑셀을 DB에 1회 적재")
    parser.add_argument("xlsx", type=Path)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    if not args.xlsx.exists():
        raise SystemExit(f"파일이 없습니다: {args.xlsx}")
    init_db()
    total, notes = import_file(args.xlsx, args.dry_run)
    mode = "검사" if args.dry_run else "적재"
    print(f"{mode} 완료: {total}개 일자")
    for note in notes:
        print("-", note)


if __name__ == "__main__":
    main()
